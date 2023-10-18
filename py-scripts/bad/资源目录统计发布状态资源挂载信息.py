import datetime
import os
import smtplib
import time
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pymongo
import pymysql
from openpyxl.workbook import Workbook

file_context = '资源目录统计发布状态资源挂载信息'
env = 'DEV'
# env = 'PROD'
receivers = None
mongo_client = None
cursor = None


region_code_to_name_dict = {
    "0500": "苏州市",
    "0505": "高新区",
    "0506": "吴中区",
    "0507": "相城区",
    "0508": "姑苏区",
    "0509": "吴江区",
    "0581": "常熟市",
    "0582": "张家港市",
    "0583": "昆山市",
    "0585": "太仓市",
    "0590": "工业园区",
    "0000": "国家级",
    "1111": "省级"
}
area_region_codes = ["0505", "0506", "0507", "0508", "0509", "0581", "0582", "0583", "0585", "0590"]


# 信息资源
info_resource_sql = "select id, info_resource_name, info_resource_provider, info_resource_life_cycle, regioncode, share_type from cmp_catalog.info_resource_platform where regioncode='%s'"

def init_envs(env):
    if env == 'PROD':
        receivers = ['jason.lu@wingconn.com', 'tommie.chen@wingconn.com', 'sisi.zhong@wingconn.com',
                     'xiao.liu@wingconn.com', 'emily.yuan@wingconn.com']
        mysql_conn = pymysql.connect(host="2.46.2.102", port=3306, user='bigdata', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient("2.46.2.239", 27017, username='admin', password='123@abcd',
                                           directConnection=True)
    else:

        receivers = ['jason.lu@wingconn.com']
        mysql_conn = pymysql.connect(host="mysql-master", port=3306, user='bigdata', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient(host="mongo-0.mongo,mongo-1.mongo,mongo-2.mongo`", port=27017)
        mongo_client.admin.authenticate('admin', '123@abcd', mechanism='SCRAM-SHA-1')
    # mysql 字典迭代器
    cursor = mysql_conn.cursor(cursor=pymysql.cursors.DictCursor)
    return mongo_client, cursor, receivers




def get_info_resource_status(status):
    if status == '0':
        return '已创建'
    elif status == '1':
        return '发布待审核'
    elif status == '2':
        return '发布已驳回'
    elif status == '3':
        return '已发布'
    elif status == '4':
        return '撤销待审核'
    elif status == '5':
        return '撤销已驳回'
    elif status == '6':
        return '已撤销'
    elif status == '7':
        return '已删除'
    return ''

# mysql查询 返回python dict array
def get_mysql_res(sql):
    cursor.execute(sql)
    res = cursor.fetchall()
    return res

# mongo查询 返回python dict array
def get_mongo_res(mongo_client, database, collection, query, projection=None):
    if projection is None:
        results = mongo_client[database][collection].find(query)
    else:
        results = mongo_client[database][collection].find(query, projection)
    return list(results)


def get_row_list(row_list, region_code, info_res_id_file_list_map, info_res_id_api_list_map, info_res_id_table_list_map):
    data_list = get_mysql_res(info_resource_sql % region_code)
    for info_resource in data_list:
        info_resource_id = str(info_resource['id'])
        file_count = len(info_res_id_file_list_map.get(info_resource_id, []))
        api_count = len(info_res_id_api_list_map.get(info_resource_id, []))
        table_count = len(info_res_id_table_list_map.get(info_resource_id, []))
        total_count = file_count + api_count + table_count
        if total_count > 0:
            has_res = '是'
        else:
            has_res = '否'

        row_list.append([
            info_resource['info_resource_name'],
            get_info_resource_status(info_resource['info_resource_life_cycle']),
            info_resource.get('share_type', ''),
            info_resource['info_resource_provider'],
            region_code_to_name_dict.get(info_resource.get('regioncode', ''), ''),
            has_res,
            total_count,
            api_count,
            file_count,
            table_count
        ])
    return row_list

def write_data_to_sheet(sheet, sheet_name, titles, data_list):
    sheet.title = sheet_name
    sheet.append(titles)
    if data_list is not None and len(data_list) > 0:
        for row in data_list:
            sheet.append(row)
    for row in sheet.iter_rows():
        for cell in row:
            cell.number_format = '@'

### 发送邮件
def send_mail(smtp_host, port, sender, password, receivers, sub, filename):
    msg = MIMEMultipart()
    msg['Subject'] = sub
    msg['From'] = sender

    part = MIMEText('本邮件为自动发送，请勿回复！')
    msg.attach(part)

    part = MIMEApplication(open(filename, 'rb').read())
    part.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(part)

    smtp = smtplib.SMTP(smtp_host, port, 'utf-8')
    smtp.login(sender, password)
    smtp.sendmail(sender, receivers, msg.as_string())
    smtp.quit()
    print('send msg successfully!')

if __name__ == '__main__':
    mongo_client, cursor, receivers = init_envs(env)
    info_res_id_api_list_map = {}
    info_res_id_table_list_map = {}
    info_res_id_file_list_map = {}
    item_resources = get_mongo_res(mongo_client, 'data_share_db', 'api_data_resource',
                                   {'status':{'$in':['published','revokeAudit','revokeRejected']}},
                                   {'_id': True, 'info_resource_id': True})
    for item in item_resources:
        key = item.get('info_resource_id', '')
        res_list = info_res_id_api_list_map.get(key, [])
        res_list.append(str(item['_id']))
        info_res_id_api_list_map[key] = res_list
    item_resources = get_mongo_res(mongo_client, 'data_share_db', 'table_data_resource',
                                   {'status':{'$in':['published','revokeAudit','revokeRejected']}},
                                   {'_id': True, 'info_resource_id': True})
    for item in item_resources:
        key = item.get('info_resource_id', '')
        res_list = info_res_id_table_list_map.get(key, [])
        res_list.append(str(item['_id']))
        info_res_id_table_list_map[key] = res_list
    item_resources = get_mongo_res(mongo_client, 'data_share_db', 'file_data_resource',
                                   {'status':{'$in':['published','revokeAudit','revokeRejected']}},
                                   {'_id': True, 'info_resource_id': True})
    for item in item_resources:
        key = item.get('info_resource_id', '')
        res_list = info_res_id_file_list_map.get(key, [])
        res_list.append(str(item['_id']))
        info_res_id_file_list_map[key] = res_list

    data_title = [
        '目录名称',
        '目录发布状态',
        '共享类型',
        '提供方',
        '提供方属地',
        '目录下是否挂接发布状态资源',
        '目录下挂接发布状态资源数量',
        '目录下挂接发布状态接口资源数量',
        '目录下挂接发布状态文件资源数量',
        '目录下挂接发布状态库表资源数量',
    ]

    filename = file_context + '.xlsx'
    # 创建一个工作簿
    wb = Workbook()
    # 市级
    write_data_to_sheet(wb.active, '市级资源目录挂接信息', data_title, get_row_list([],'0500', info_res_id_file_list_map, info_res_id_api_list_map, info_res_id_table_list_map))
    # 区县
    area_row_list = []
    for area_region_code in area_region_codes:
        area_row_list = get_row_list(area_row_list,area_region_code, info_res_id_file_list_map, info_res_id_api_list_map, info_res_id_table_list_map)
    write_data_to_sheet(wb.create_sheet(), '区县资源目录挂接信息', data_title, area_row_list)
    # 将工作簿保存到 xlsx 文件
    wb.save(filename)

    ###
    sub = Header('[' + env + ']-[' + file_context + ']' + str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                 'utf-8')
    send_mail('smtp.163.com', 25, 'rikurobot@163.com', 'BDDHTVARWQRQFPFN', receivers, sub, filename)

    # 删除文件
    os.remove(filename)
