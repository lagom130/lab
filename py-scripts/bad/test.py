from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment

data = [
    {"Name": "Alice", "Age": 25, "City": "New York", "Labels": [{"name": 1, "age": 1}, {"name": 1, "age": 1}]},
    {"Name": "Bob", "Age": 30, "City": "Los Angeles", "Labels": [{"name": 1, "age": 1}, {"name": 1, "age": 1}]},
    {"Name": "Carol", "Age": 35, "City": "Chicago", "Labels": [{"name": 1, "age": 1}, {"name": 1, "age": 1}]}
]

# 创建一个新的工作簿
wb = Workbook()
sheet = wb.active

# 写入表头
headers = list(data[0].keys())
header_row = 1
for col_index, header in enumerate(headers, start=1):
    sheet.cell(row=header_row, column=col_index, value=header)

# 写入数据
data_row = 2
for item in data:
    for col_index, key in enumerate(headers, start=1):
        value = item[key]
        cell = sheet.cell(row=data_row, column=col_index, value=value)

        # 设置单元格对齐方式
        if key == "City":
            cell.alignment = Alignment(horizontal="center")

    data_row += 1

# 合并单元格
city_column = sheet['C']
previous_city = None
start_row = None

for cell in city_column:
    if cell.row == 1:
        previous_city = cell.value
        start_row = cell.row
        continue

    current_city = cell.value

    if current_city != previous_city:
        end_row = cell.row - 1

        # 合并单元格
        if start_row != end_row:
            sheet.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)

        previous_city = current_city
        start_row = cell.row

# 保存工作簿
wb.save('your_file.xlsx')