import csv
import json
import os
import datetime
from email.header import Header

from openpyxl.workbook import Workbook

import base

# env = '326'
# env = 'DEV'
env = 'DEV'
tsv_path = "E:\\开放平台信息资源目录（全量）.tsv"
receivers = None
mongo_client = None
cursor = None
file_context = '开放目录比对结果'

def readCsvDataToDict(fp):
    data_dict = {}
    with open(fp, 'r', encoding="UTF-8") as f:
        reader = csv.DictReader(f, delimiter='\t')
        for row in reader:
            info_item = {
                "information_item_name":row['information_item_name'],
                "item_open_attribute":row['item_open_attribute'],
                "sensitive_level":row['sensitive_level'],
            }
            resource_code = row['resource_code']
            open_resource = data_dict.get(resource_code, {
                'open_id': row['id'],
                'resource_code': row['resource_code'],
                'resource_name': row['resource_name'],
                'catalog_open_attribute': row['catalog_open_attribute'],
                'info_items':[],
                'resource_status':row['resource_status']
            })
            open_resource['info_items'].append(info_item)
            data_dict[resource_code] = open_resource

    return data_dict


def is_diff(open_resource, info_resource):
    if info_resource is None:
        return True, "共享平台目录不存在", ''
    open_info_items = open_resource['info_items']
    info_resource_info_items = {}
    try:
        info_resource_info_items = json.loads(info_resource.get('info_item_desc_detail', ''))
    except:
        print(info_resource['info_resource_name'] + " info item detail json loads error, detail:" + str(
            info_resource.get('info_item_desc_detail', '')))
    for open_info_item in open_info_items:
        info_item_name = open_info_item['information_item_name']
        info_resource_info_item = next((info_item for info_item in info_resource_info_items if
                         info_item.get('infoItemName', '') == info_item_name),None)
        if info_resource_info_item is None:
            return True, "共享平台未找到信息项", info_item_name
        info_resource_info_item_is_open = False
        if info_resource_info_item.get('isOpen', "0") == "1":
            info_resource_info_item_is_open = True
        if open_info_item['item_open_attribute'] != '不予开放' and info_resource_info_item_is_open is False:
            return True, "信息项是否开放不一致",info_item_name+"在共享平台isOpen="+str(info_resource_info_item_is_open)+", 在开放平台为"+open_info_item['item_open_attribute']
        if open_info_item['item_open_attribute'] == '不予开放' and info_resource_info_item_is_open is True:
            return True, "信息项是否开放不一致",info_item_name+"在共享平台isOpen="+str(info_resource_info_item_is_open)+", 在开放平台为"+open_info_item['item_open_attribute']
        if open_info_item['sensitive_level'] is not None and open_info_item['sensitive_level'] == 'NULL':
            open_info_item['sensitive_level'] = None
        # if info_resource_info_item.get('sensitiveLevel', None) != open_info_item['sensitive_level']:
        #     return True,"敏感级别不一致",info_item_name+": 开放为"+str(info_resource_info_item.get('sensitiveLevel', None))+",共享为"+str(open_info_item['sensitive_level'])

    info_resource_open_attribute = '不予开放'
    if info_resource['info_resource_life_cycle'] not in ['3', '4', '5']:
        return True, "共享平台非发布状态", "info_resource_life_cycle=" + str(info_resource['info_resource_life_cycle'])
    if info_resource['is_open'] == '1':
        if info_resource['open_type'] == 'unconditional':
            info_resource_open_attribute = '无条件开放'
        elif info_resource['open_type'] == 'conditional':
            info_resource_open_attribute = '有条件开放'
    # 目录开放类型判断
    if info_resource_open_attribute != open_resource['catalog_open_attribute']:
        return True, "目录开放类型不一致", "共享平台为" + info_resource_open_attribute + ', 开放平台为' + open_resource[
            'catalog_open_attribute']
    if info_resource['info_resource_name'] != open_resource['resource_name']:
        return True, '目录名称不一致', 'info_resource_name=' + info_resource[
            'info_resource_name'] + ', open_resource_name=' + open_resource['resource_name']
    return False, None, None


if __name__ == '__main__':
    mongo_client, cursor, receivers, mysql_conn = base.init_envs(env)
    info_resources = base.get_mysql_res(cursor, "select id, info_resource_name, info_resource_provider, base_info_resource_code, theme_info_resource_code, dept_info_resource_code, info_resource_life_cycle, is_open, open_type, info_item_desc_detail from cmp_catalog.info_resource_platform WHERE regioncode IN ('0500','0505','0506','0507','0508','0509','0581','0582','0583','0585','0590')")
    data_dict = readCsvDataToDict(tsv_path)
    # 不同的列表
    diff_info_resource_ids = []
    diff_info_resources =[]
    remarks = []
    # 找不到的列表
    not_found_open_resource_ids = []
    not_found_open_resources = []
    for resource_code in data_dict.keys():
        open_resource = data_dict.get(resource_code)
        info_resources_by_name = [info_resource for info_resource in info_resources if info_resource.get('info_resource_name') == open_resource['resource_name']]
        info_resource = next((info_resource for info_resource in info_resources_by_name if info_resource.get('dept_info_resource_code') == resource_code or info_resource.get('base_info_resource_code') == resource_code or info_resource.get('theme_info_resource_code') == resource_code), None)
        if info_resource is None:
            if len(info_resources_by_name) == 1:
                info_resource = info_resources_by_name[0]
                remarks.append([open_resource['open_id'], open_resource['resource_status'], resource_code, open_resource['resource_name'], '编码不正确', info_resource['id']])
            if len(info_resources_by_name)>1:
                remarks.append([open_resource['open_id'], open_resource['resource_status'], resource_code, open_resource['resource_name'], '编码不正确,同名对应多个目录', str([info_resource['id'] for info_resource in info_resources_by_name])])
            if len(info_resources_by_name)==0:
                info_resources_by_code = [info_resource for info_resource in info_resources if info_resource.get(
                    'dept_info_resource_code') == resource_code or info_resource.get(
                    'base_info_resource_code') == resource_code or info_resource.get(
                    'theme_info_resource_code') == resource_code]
                if len(info_resources_by_code) == 1:
                    info_resource = info_resources_by_code[0]
                if len(info_resources_by_code) == 0:
                    not_found_open_resource_ids.append(open_resource['open_id'])
                    not_found_open_resources.append([open_resource['open_id'], open_resource['resource_status'], open_resource['resource_name'], str(open_resource)])
                if len(info_resources_by_code) > 1:
                    remarks.append([open_resource['open_id'], open_resource['resource_status'], resource_code, open_resource['resource_name'], '名称不正确,编码对应多个目录', str([info_resource['id'] for info_resource in info_resources_by_code])])
        if info_resource is not None:
            flag, result_type, result_detail = is_diff(open_resource, info_resource)
            if flag:
                if info_resource is None:
                    not_found_open_resource_ids.append(open_resource['open_id'])
                    not_found_open_resources.append([open_resource['open_id'], open_resource['resource_status'], str(open_resource)])
                else:
                    diff_info_resource_ids.append(info_resource['id'])
                    diff_info_resources.append([open_resource['open_id'], open_resource['resource_status'],
                                                info_resource['id'], result_type, result_detail,
                                                open_resource.get('catalog_open_attribute'),
                                                str(info_resource.get('info_resource_life_cycle')),
                                                str(info_resource.get('is_open')),
                                                str(info_resource.get('open_type')),
                                                str(open_resource['info_items']),
                                                str(info_resource['info_item_desc_detail'])])

    filename = file_context + '.xlsx'
    # 创建一个工作簿
    wb = Workbook()

    # 获取活动工作表
    ws1 = wb.active
    ws1.append(['开放目录数量', str(len(data_dict.keys()))])
    ws1.append(['不同目录ID', str(len(diff_info_resource_ids)), str(diff_info_resource_ids)])
    ws1.append(['找不到目录开放ID', str(len(not_found_open_resource_ids)), str(not_found_open_resource_ids)])
    for remark in remarks:
        ws1.append(remark)
    # 将所有单元格格式设置为字符串
    for row in ws1.iter_rows():
        for cell in row:
            cell.number_format = '@'
    ws2 = wb.create_sheet('不同目录')
    ws2.append(['开放平台ID','开放平台状态','共享平台目录ID', '原因', '开放平台开放类型','共享平台生命周期','共享平台是否开放','共享平台开放类型','开放平台信息项', '共享平台信息项'])
    # 将数据写入工作表
    for row in diff_info_resources:
        ws2.append(row)
    # 将所有单元格格式设置为字符串
    for row in ws2.iter_rows():
        for cell in row:
            cell.number_format = '@'
    ws3 = wb.create_sheet('找不到的目录')
    ws3.append(['开放平台ID','开放平台状态','开放平台目录信息'])

    # 将数据写入工作表
    for row in not_found_open_resources:
        ws3.append(row)
    # 将所有单元格格式设置为字符串
    for row in ws3.iter_rows():
        for cell in row:
            cell.number_format = '@'
    # 将工作簿保存到 xlsx 文件
    wb.save(filename)

    ###
    sub = Header('[' + env + ']-[' + file_context + ']' + str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                 'utf-8')
    base.send_mail('smtp.163.com', 25, 'rikurobot@163.com', 'BDDHTVARWQRQFPFN', ['jason.lu@wingconn.com'], sub, filename)

    # 删除文件
    os.remove(filename)
    print('export task ended at:' + datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
