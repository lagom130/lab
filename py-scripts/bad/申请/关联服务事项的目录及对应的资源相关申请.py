import csv
import datetime
import gc
import itertools
import json
import os
import re
import smtplib
import time
import zipfile
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pymongo
import pymysql
from openpyxl.workbook import Workbook

file_context = '关联服务事项的目录及对应的资源相关申请单'
env = '326'
# env = 'DEV'
# env = 'PROD'
receivers = None
mongo_client = None
cursor = None

###
context = '一件事'

#########################################################################################################
######## 字典 ###########################################################################################
# 属地字典
region_code_to_name_dict = {
    "0500": "苏州市",
    "0505": "高新区",
    "0506": "吴中区",
    "0507": "相城区",
    "0508": "姑苏区",
    "0509": "吴江区",
    "0581": "常熟市",
    "0582": "张家港市",
    "0583": "昆山市",
    "0585": "太仓市",
    "0590": "工业园区",
    "0000": "国家级",
    "1111": "省级"
}
city_region = '0500'
area_regions = ['0505', '0506', '0507', '0508', '0509', '0581', '0582', '0583', '0585', '0590']
province_regions = ['0000', '1111']
filter_info_resource_code_prefix = '30701505003030050000000003'
# 目录分类（按行业）
belong_industry_dict = {
    "A": "农、林、牧、渔业",
    "B": "采矿业",
    "C": "制造业",
    "D": "电力、势力、燃气及水生产和供应业",
    "E": "建筑业",
    "F": "批发和零售业",
    "G": "交通运输、仓储和邮政业",
    "H": "住宿和餐饮业",
    "I": "信息传输、软件和信息技术服务业",
    "J": "金融业",
    "K": "房地产业",
    "L": "租赁和商务服务业",
    "M": "科学研究和技术服务业",
    "N": "水利、环境和公共设施管理业",
    "O": "居民服务、修理和其他服务业",
    "P": "教育",
    "Q": "卫生和社会工作",
    "R": "文化、体育和娱乐业",
    "S": "公共管理、社会保障和社会组织",
    "T": "国际组织"
}
# 布尔中文-字符串01
bool_number_string_dict = {
    "1": "是",
    "0": "否"
}
# 是否电子证照
certification_type_dict = {
    "1": "是",
    "2": "否"
}
# 事项目录所属领域
field_type_dict = {
    "1": "科技创新",
    "2": "商贸流通",
    "3": "社会救助",
    "4": "城建住房",
    "5": "教育文化",
    "6": "工业农业",
    "7": "机构团体",
    "8": "地理空间",
    "9": "资源能源",
    "10": "市场监管",
    "11": "生活服务",
    "12": "生态环境",
    "13": "交通运输",
    "14": "安全生产",
    "15": "社保就业",
    "16": "医疗卫生",
    "17": "信用服务",
    "18": "公共服务",
    "19": "财税金融",
    "20": "气象服务",
    "21": "法律服务",
    "22": "新冠疫苗",
    "23": "其他"
}
# 应用场景
gov_use_scene_dict = {
    "1": "政务服务",
    "2": "公共服务",
    "3": "监管",
    "0": "其他"
}
# 目录生命周期
info_res_life_cycle_dict = {
    "0": "已创建",
    "1": "发布待审核",
    "2": "发布已驳回",
    "3": "已发布",
    "4": "撤销待审核",
    "5": "撤销已驳回",
    "6": "已撤销",
    "7": "已删除"
}
# 信息项共享类型
item_share_type_dict = {
    "1": "无条件共享",
    "2": "有条件共享",
    "3": "不予共享"
}
# 提供渠道
net_type_dict = {
    "1": "政务外网",
    "2": "互联网",
    "3": "部门专网"
}
# 开放类型
open_type_dict = {
    "unconditional": "无条件开放",
    "conditional": "有条件开放",
    "disallowed": "不开放"
}
# 不xx依据
reason_type_dict = {
    1: "法律法规",
    2: "政策文件",
    3: "涉及敏感信息",
    4: "其他"
}
# 所属领域
resource_subject_dict = {
    "01": "婚育收养",
    "02": "环境资源",
    "03": "教育培训",
    "04": "住房保障",
    "05": "劳动就业",
    "06": "执业资格",
    "07": "社会保障",
    "08": "纳税缴费",
    "09": "医疗卫生",
    "10": "交通旅游",
    "11": "出境入境",
    "12": "生活服务",
    "13": "民族宗教",
    "14": "三农服务",
    "15": "死亡殡葬",
    "16": "设立变更",
    "17": "行业准营",
    "18": "投资立项",
    "19": "工程建设",
    "20": "涉外服务",
    "21": "安全生产",
    "22": "破产注销",
    "23": "文物保护",
    "24": "其他"
}
# 数据敏感级别
sensitive_level_dict = {
    "1": "1级",
    "2": "2级",
    "3": "3级",
    "4": "4级"
}
# 存储类型
storage_mode_dict = {
    "API": "接口",
    "TABLE": "数据库",
    "FILE_STRUCTURED": "结构化文件",
    "FILE_UNSTRUCTURED": "非结构化文件"
}
# 事项类型
task_type_dict = {
    "GOV": "政务服务事项",
    "PUBLIC": "公共服务事项"
}
# 目录分类（按主题）
use_scene_dict = {
    "01": "社会保险",
    "02": "交通出行",
    "03": "婚育",
    "04": "社区周边生活服务",
    "05": "政府办事",
    "06": "就医与保健",
    "07": "学校教育与终身教育",
    "08": "培训与就业",
    "09": "城市安全"
}
# 布尔值字典
bool_dict = {
    True: '是',
    False: '否'
}
# 布尔值字符串字典
bool_str_dict = {
    'true': '是',
    'false': '否'
}
# 布尔值字符串字典
bool_yn_dict = {
    'Y': '是',
    'N': '否'
}
# 更新周期
update_cycle_dict = {
    'year': '每年',
    'quarter': '季度',
    'month': '每月',
    'week': '每周',
    'day': '每日',
    'ondemand': '按需',
    '实时': '实时',
    '每日': '每日',
    '每周': '每周',
    '每月': '每月',
    '季度': '季度',
    '每年': '每年',
    '按需': '按需'
}

# 资源状态
data_res_status_dict = {
    'created': '已创建',
    'publishAudit': '发布待审核',
    'published': '已发布',
    'publishRejected': '发布已驳回',
    'revokeAudit': '撤销待审核',
    'revoked': '已撤销',
    'revokeRejected': '撤销已驳回',
    'closed': '已关闭',
}
# 匹配模式
url_map_mode_dict = {
    'prefix': '前缀匹配',
    'absolute': '绝对匹配',
}
# 请求类型
api_protocol_dict = {
    'rs': 'restful',
    'ws': 'webservice'
}
# 资源类型
resource_type_dict = {
    'API': '接口',
    'TABLE': '库表',
    'FILE': '文件',
    'api': '接口',
    'table': '库表',
    'file': '文件',
}
# 申请状态
apply_status_dict = {
    'created': '已创建',
    'unAudit': '待审核',
    'auditReject': '已退回',
    'unSubscrib': '待订阅',
    'unSend': '待推送',
    'finished': '已完成',
    'invalid': '失效',
    'handOver': '转交'
}
# 省级上报状态
report_status_dict = {
    None: '未发布',
    '': '未发布',
    'handled': '已绑定',
    'initializing': '初始化中',
    'initialized': '初始化完成',
    'migrating': '迁移中',
    'migrated': '迁移完成',
    'published': '已发布',
    'unpublished': '不发布'}
# 目录省级上报状态
info_resource_report_status_dict = {
    None: '未上报',
    '': '未上报',
    'success': '已上报'
}

###########################################################################################################
######## 通用方法 ###########################################################################################

# 根据字典key获取字典值，缺省默认值为空字符串
def dict_key_to_value(dictionary, key, default_value=""):
    return dictionary.get(key, default_value)


# 根据字典值获取字典key，缺省默认值为空字符串
def dict_value_to_key(dictionary, value, default_value=""):
    for k, v in dictionary.items():
        if v == value:
            return k
    return default_value


def get_element(arr, i, default_value=""):
    try:
        return arr[i]
    except IndexError:
        return default_value


# 单一实体占用多行写入单元格，只有第一行会写
def write_cell(entity_index, value):
    if entity_index == 0:
        return value
    return None


# 单一实体占用多行写入单元格，只有第一行会写，并将值翻译
def write_cell_convert_by_dict(entity_index, dictionary, key, default_value=""):
    if entity_index == 0:
        return dict_key_to_value(dictionary, key, default_value)
    return None


# 单一实体占用多行写入单元格，只有第一行会写
def entity_to_cell(entity_index, entity, key):
    if entity_index == 0:
        return entity.get(key, None)
    return None


# 单一实体占用多行写入单元格，只有第一行会写，并将值翻译
def entity_to_cell_convert_by_dict(entity_index, dictionary, entity, key, default_value=""):
    if entity_index == 0:
        return dict_key_to_value(dictionary, entity.get(key, default_value), default_value)
    return None


# 初始化环境信息
def init_envs(env):
    if env == 'PROD':
        receivers = ['jason.lu@wingconn.com', 'tommie.chen@wingconn.com', 'sisi.zhong@wingconn.com', 'xiao.liu@wingconn.com','emily.yuan@wingconn.com']
        mysql_conn = pymysql.connect(host="2.46.2.102", port=3306, user='bigdata', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient("2.46.2.239", 27017, username='admin', password='123@abcd',
                                           directConnection=True)
    elif env == '326':
        receivers = ['jason.lu@wingconn.com']
        mysql_conn = pymysql.connect(host="10.10.32.6", port=3306, user='wingtest', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient("10.10.32.6", port=27017)
        mongo_client.admin.authenticate('admin', '123@abcd', mechanism='SCRAM-SHA-1')
    else:

        receivers = ['jason.lu@wingconn.com']
        mysql_conn = pymysql.connect(host="mysql-master", port=3306, user='bigdata', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient(host="mongo-0.mongo,mongo-1.mongo,mongo-2.mongo`", port=27017)
        mongo_client.admin.authenticate('admin', '123@abcd', mechanism='SCRAM-SHA-1')
    # mysql 字典迭代器
    cursor = mysql_conn.cursor(cursor=pymysql.cursors.DictCursor)
    return mongo_client, cursor, receivers


# 13位时间戳转格式化日期字符串
def timestamp_to_date_str(time_stamp, format_string="%Y-%m-%d %H:%M:%S"):
    try:
        if time_stamp is None or time_stamp == '':
            return ''
        if len(str(time_stamp)) > 13:
            time_stamp = int(str(time_stamp)[:13])
        elif len(str(time_stamp)) < 13:
            time_stamp = int('{}{}'.format(str(time_stamp), '0' * (10 - len(str(time_stamp)))))

        time_array = time.localtime((time_stamp / 1000))
        return time.strftime(format_string, time_array)
    except BaseException as e:
        print(str(time_stamp))
        return ''


# mysql查询 返回python dict array
def get_mysql_res(sql):
    cursor.execute(sql)
    res = cursor.fetchall()
    return res


# mongo查询 返回python dict array
def get_mongo_res(mongo_client, database, collection, query, projection=None):
    if projection is None:
        results = mongo_client[database][collection].find(query)
    else:
        results = mongo_client[database][collection].find(query, projection)
    return list(results)


def write_data_to_sheet(sheet, sheet_name, titles, data_list):
    sheet.title = sheet_name
    sheet.append(titles)
    if data_list is not None and len(data_list) > 0:
        for row in data_list:
            sheet.append(row)
    for row in sheet.iter_rows():
        for cell in row:
            cell.number_format = '@'


def write_csv_to_zip(path, title, data_list, zip_file):
    with open(path, 'w', encoding='utf-8', newline='') as cf:
        writer = csv.writer(cf)
        writer.writerow(title)
        if data_list is not None:
            for item in data_list:
                for cell in item:
                    if cell is None or cell is 'null':
                        cell = ''
                writer.writerow(item)
    zip_file.write(path)
    os.remove(path)


### 发送邮件
def send_mail(smtp_host, port, sender, password, receivers, sub, filename):
    msg = MIMEMultipart()
    msg['Subject'] = sub
    msg['From'] = sender

    part = MIMEText('本邮件为自动发送，请勿回复！')
    msg.attach(part)

    part = MIMEApplication(open(filename, 'rb').read())
    part.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(part)

    smtp = smtplib.SMTP(smtp_host, port, 'utf-8')
    smtp.login(sender, password)
    smtp.sendmail(sender, receivers, msg.as_string())
    smtp.quit()
    print('send msg successfully!')


# 列表转字符串（sql拼接用）
def get_arr_str(arr):
    return "'" + "','".join(arr) + "'"

if __name__ == '__main__':
    mongo_client, cursor, receivers = init_envs(env)
    info_resources = get_mysql_res("select id, info_resource_name, info_resource_provider, task_name from cmp_catalog.info_resource_platform where related_task_flag = 1 or (task_name is not null and task_name != '')")
    info_res_ids = []
    info_res_dict = dict()
    for info_res in info_resources:
        info_res_ids.append(info_res['id'])
        info_res_dict[info_res['id']] = info_res
    row_list = []
    resources = get_mysql_res("select id, info_resource_id, info_resource_name, info_resource_provider from data_share_db.share_data_resource where info_resource_id in (%s)" % get_arr_str(info_res_ids))
    share_data_resource_ids = []
    res_dict = {}
    for res in resources:
        share_data_resource_ids.append(res['id'])
        res_dict[res['id']] = res
    applies = get_mongo_res(mongo_client, 'data_share_db', 'data_resource_apply',
                            {'share_resource_id':{'$in': share_data_resource_ids}},
                            {'_id': 1, 'resource_id': 1, 'share_resource_id': 1, 'resource_name': 1, 'apply_type':1,'code': 1, 'status': 1,
                             'dep_name': 1, 'business_sys_info': 1, 'origin_incident': 1,'use_range':1, 'apply_time': 1,'data_creater_dep_name':1})



    row_list = []
    row_list.append(['申请主键','resourceId','shareResourceId','目录主键','所属目录','提供方','资源名称','资源类型','申请流水号','申请状态','申请部门','业务系统名称','业务系统IP','系统部署地址','申请事由','申请使用范围','申请时间','事项名称'])
    for apply in applies:
        res = res_dict.get(apply.get('share_resource_id', ''), {})
        info_res = info_res_dict.get(res.get('info_resource_id', ''), {})
        resource_id = apply.get('resource_id','')
        business_sys_info = apply.get('business_sys_info', {})
        #业务系统名称
        sys_name = business_sys_info.get('sys_name', '')
        #业务系统ip
        sys_ip = business_sys_info.get('sys_ip', '')
        #业务系统部署地点
        sys_location = business_sys_info.get('sys_location', '')
        row_list.append([
            str(apply['_id']),
            str(resource_id),
            apply.get('share_resource_id', ''),
            info_res.get('id', ''),
            info_res.get('info_resource_name', ''),
            info_res.get('info_resource_provider', ''),
            apply.get('resource_name', ''),
            apply.get('apply_type', ''),
            apply.get('code', ''),
            dict_key_to_value(apply_status_dict,apply.get('status', ''),),
            apply.get('dep_name', ''),
            sys_name,
            sys_ip,
            sys_location,
            apply.get('origin_incident', ''),
            apply.get('use_range', ''),
            timestamp_to_date_str(apply.get('apply_time', '')),
            info_res.get('task_name', '')
        ])

    filename = file_context + '.xlsx'
    # 创建一个工作簿
    wb = Workbook()

    # 获取活动工作表
    ws = wb.active
    ws.title = file_context
    # 将数据写入工作表
    for row in row_list:
        ws.append(row)
        print(row)
    # 将所有单元格格式设置为字符串
    for row in ws.iter_rows():
        for cell in row:
            cell.number_format = '@'


    # 将工作簿保存到 xlsx 文件
    wb.save(filename)

    ###
    sub = Header('[' + env + ']-[' + file_context + ']' + str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                 'utf-8')
    send_mail('smtp.163.com', 25, 'rikurobot@163.com', 'BDDHTVARWQRQFPFN', receivers, sub, filename)

    # 删除文件
    os.remove(filename)

    print('export task ended at:' + datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
