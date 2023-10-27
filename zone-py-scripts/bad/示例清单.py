import datetime
import os
import smtplib
import time
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pymongo
import pymysql
from openpyxl.workbook import Workbook

file_context = '示例清单'
env = 'DEV'
# env = 'PROD'
receivers = None
mongo_client = None
cursor = None

region_code_to_name_dict = {
    "0500": "苏州市",
    "0505": "高新区",
    "0506": "吴中区",
    "0507": "相城区",
    "0508": "姑苏区",
    "0509": "吴江区",
    "0581": "常熟市",
    "0582": "张家港市",
    "0583": "昆山市",
    "0585": "太仓市",
    "0590": "工业园区",
    "0000": "国家级",
    "1111": "省级"
}


def get_info_resource_status(status):
    if status == '0':
        return '已创建'
    elif status == '1':
        return '发布待审核'
    elif status == '2':
        return '发布已驳回'
    elif status == '3':
        return '已发布'
    elif status == '4':
        return '撤销待审核'
    elif status == '5':
        return '撤销已驳回'
    elif status == '6':
        return '已撤销'
    elif status == '7':
        return '已删除'
    return ''


# 共享资源状态转换
def get_share_data_resource_status(status):
    if status == 'created':
        return '已创建'
    elif status == 'publishAudit':
        return '发布待审核'
    elif status == 'published':
        return '已发布'
    elif status == 'publishRejected':
        return '发布已驳回'
    elif status == 'revokeAudit':
        return '撤销待审核'
    elif status == 'revoked':
        return '已撤销'
    elif status == 'revokeRejected':
        return '撤销已驳回'
    elif status == 'closed':
        return '已关闭'
    return ''


def init_envs(env):
    if env == 'PROD':
        receivers = ['jason.lu@wingconn.com', 'tommie.chen@wingconn.com', 'sisi.zhong@wingconn.com',
                     'xiao.liu@wingconn.com', 'emily.yuan@wingconn.com']
        mysql_conn = pymysql.connect(host="2.46.2.102", port=3306, user='bigdata', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient("2.46.2.239", 27017, username='admin', password='123@abcd',
                                           directConnection=True)
    else:

        receivers = ['jason.lu@wingconn.com']
        mysql_conn = pymysql.connect(host="mysql-master", port=3306, user='bigdata', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient(host="mongo-0.mongo,mongo-1.mongo,mongo-2.mongo`", port=27017)
        mongo_client.admin.authenticate('admin', '123@abcd', mechanism='SCRAM-SHA-1')
    # mysql 字典迭代器
    cursor = mysql_conn.cursor(cursor=pymysql.cursors.DictCursor)
    return mongo_client, cursor, receivers


# mysql查询 返回python dict array
def get_mysql_res(sql):
    cursor.execute(sql)
    res = cursor.fetchall()
    return res


# mongo查询 返回python dict array
def get_mongo_res(mongo_client, database, collection, query, projection=None):
    if projection is None:
        results = mongo_client[database][collection].find(query)
    else:
        results = mongo_client[database][collection].find(query, projection)
    return list(results)


### 发送邮件
def send_mail(smtp_host, port, sender, password, receivers, sub, filename):
    msg = MIMEMultipart()
    msg['Subject'] = sub
    msg['From'] = sender

    part = MIMEText('本邮件为自动发送，请勿回复！')
    msg.attach(part)

    part = MIMEApplication(open(filename, 'rb').read())
    part.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(part)

    smtp = smtplib.SMTP(smtp_host, port, 'utf-8')
    smtp.login(sender, password)
    smtp.sendmail(sender, receivers, msg.as_string())
    smtp.quit()
    print('send msg successfully!')


def compare_str_arrays(arr1, arr2):
    return set(arr1) == set(arr2)

# 13位时间戳转格式化日期字符串
def timestamp_to_date_str(time_stamp, format_string="%Y-%m-%d %H:%M:%S"):
    try:
        if time_stamp is None or time_stamp == '':
            return ''
        if len(str(time_stamp)) > 13:
            time_stamp = int(str(time_stamp)[:13])
        elif len(str(time_stamp)) < 13:
            time_stamp = int('{}{}'.format(str(time_stamp), '0'*(10-len(str(time_stamp)))))

        time_array = time.localtime((time_stamp / 1000))
        return time.strftime(format_string, time_array)
    except BaseException as e:
        print(str(time_stamp))
        return ''

def info_res_to_row_list(data_list, info_resource_provider_dict):
    info_res_row_list = []
    for item in data_list:
        info_res_row_list.append([
            item.get('id', ''),
            item.get('info_resource_name', ''),
            item.get('info_resource_provider', ''),
            region_code_to_name_dict.get(item['regioncode'], ''),
            get_info_resource_status(item.get('info_resource_life_cycle', '')),
            timestamp_to_date_str(item.get('public_date', '')),
            item.get('share_type', ''),
        ])
        info_resource_provider_dict[item['id']] = item['info_resource_provider']
    return info_res_row_list, info_resource_provider_dict


def data_res_to_row_list(data_list_dict, region_codes):
    data_res_row_list = []
    for rc in region_codes:
        data_list = data_list_dict.get(rc, [])
        for item in data_list:
            data_res_row_list.append([
                item.get('id', ''),
                item.get('resource_name', ''),
                item.get('resource_type', ''),
                item.get('creater_dep_name', ''),
                region_code_to_name_dict.get(rc, ''),
                item.get('status', ''),
                item.get('publish_time',''),
                item.get('share_range',''),
            ])
    return data_res_row_list


def write_data_to_sheet(sheet, sheet_name, titles, data_list):
    sheet.title = sheet_name
    sheet.append(titles)
    if data_list is not None and len(data_list) > 0:
        for row in data_list:
            sheet.append(row)
    for row in sheet.iter_rows():
        for cell in row:
            cell.number_format = '@'


if __name__ == '__main__':
    mongo_client, cursor, receivers = init_envs(env)
    sql = "select id, info_resource_name, info_resource_provider, regioncode,public_date, info_resource_life_cycle, share_type from cmp_catalog.info_resource_platform";
    info_res_title = ['主键', '目录名称', '提供方', '提供方属地', '状态', '发布时间', '共享类型']
    share_data_res_title = ['主键', '资源名称', '资源类型', '提供方', '提供方属地', '状态', '发布时间', '共享类型']
    info_resource_provider_dict = {}

    filename = file_context + '.xlsx'
    # 创建一个工作簿
    wb = Workbook()

    # # sheet1 市级示例目录清单
    # print('>>>>>>>>>>>>>> start export 市级示例目录清单')
    # city_info_res_list, info_resource_provider_dict = info_res_to_row_list(
    #     get_mysql_res(sql + " where regioncode='0500'"), info_resource_provider_dict)
    # write_data_to_sheet(wb.active, '市级示例目录清单', info_res_title, city_info_res_list)
    # # sheet2 区县示例目录清单
    # print('>>>>>>>>>>>>>> start export 区县示例目录清单')
    # area_info_res_list, info_resource_provider_dict = info_res_to_row_list(
    #     get_mysql_res(
    #         sql + " where regioncode in ('0505','0506','0507','0508','0509','0581','0582','0583','0585','0590')"),
    #     info_resource_provider_dict)
    # write_data_to_sheet(wb.create_sheet(), '区县示例目录清单', info_res_title, area_info_res_list)

    # 资源预处理
    region_to_share_data_res_list_dict = {}
    files = get_mongo_res(mongo_client, 'data_share_db', 'file_data_resource',
                          {'related_catalog_info.region_code': {'$nin': ['1111', '0000']}},
                          {'_id': True, 'info_resource_id': True, 'creater_dep_name': True,
                           'related_catalog_info': True, 'status': True, 'publish_time': True, 'share_range': True,
                           'file_name': True})
    for item in files:
        region_code = item.get('related_catalog_info', {}).get('region_code', '')
        res_list = region_to_share_data_res_list_dict.get(region_code, [])
        res_list.append({
            'id': str(item.get('_id', '')),
            'resource_name': item.get('file_name', ''),
            'resource_type': '文件',
            'creater_dep_name': item.get('creater_dep_name', ''),
            'region_code': region_code,
            'status': get_share_data_resource_status(item.get('status', '')),
            'publish_time': item.get('publish_time', ''),
            'share_range': item.get('share_range', '')
        })
        region_to_share_data_res_list_dict[region_code] = res_list
    files = None

    tables = get_mongo_res(mongo_client, 'data_share_db', 'table_data_resource',
                           {'related_catalog_info.region_code': {'$nin': ['1111', '0000']}},
                           {'_id': True, 'info_resource_id': True, 'creater_dep_name': True,
                            'related_catalog_info': True, 'status': True, 'publish_time': True, 'share_range': True,
                            'table_name': True})
    for item in tables:
        region_code = item.get('related_catalog_info', {}).get('region_code', '')
        res_list = region_to_share_data_res_list_dict.get(region_code, [])
        res_list.append({
            'id': str(item.get('_id', '')),
            'resource_name': item.get('table_name', ''),
            'resource_type': '库表',
            'creater_dep_name': item.get('creater_dep_name', ''),
            'region_code': region_code,
            'status': get_share_data_resource_status(item.get('status')),
            'publish_time': item.get('publish_time', ''),
            'share_range': item.get('share_range', '')
        })
        region_to_share_data_res_list_dict[region_code] = res_list
    tables = None

    apis = get_mongo_res(mongo_client, 'data_share_db', 'table_data_resource',
                         {'related_catalog_info.region_code': {'$nin': ['1111', '0000']}},
                         {'_id': True, 'info_resource_id': True, 'creater_dep_name': True,
                          'related_catalog_info': True, 'status': True, 'publish_time': True, 'share_range': True,
                          'api_name': True})
    for item in apis:
        region_code = item.get('related_catalog_info', {}).get('region_code', '')
        res_list = region_to_share_data_res_list_dict.get(region_code, [])
        res_list.append({
            'id': str(item.get('_id', '')),
            'resource_name': item.get('table_name', ''),
            'resource_type': '库表',
            'creater_dep_name': item.get('creater_dep_name', ''),
            'region_code': region_code,
            'status': get_share_data_resource_status(item.get('status', '')),
            'publish_time': item.get('publish_time', ''),
            'share_range': item.get('share_range', '')
        })
        region_to_share_data_res_list_dict[region_code] = res_list
    apis = None

    # sheet3 市级资源清单
    print('>>>>>>>>>>>>>> start export 市级示例资源清单')
    write_data_to_sheet(wb.create_sheet(), '市级示例资源清单', share_data_res_title,
                        data_res_to_row_list(region_to_share_data_res_list_dict, ['0500']))
    # sheet4 市级资源清单
    print('>>>>>>>>>>>>>> start export 区县示例资源清单')
    write_data_to_sheet(wb.create_sheet(), '区县示例资源清单', share_data_res_title,
                        data_res_to_row_list(region_to_share_data_res_list_dict,
                                             ['0505', '0506', '0507', '0508', '0509', '0581', '0582', '0583', '0585',
                                              '0590']))

    # 将工作簿保存到 xlsx 文件
    wb.save(filename)
    print('>>>>>>>>>>>>>> excel save')

    ###
    sub = Header('[' + env + ']-[' + file_context + ']' + str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                 'utf-8')
    send_mail('smtp.163.com', 25, 'rikurobot@163.com', 'BDDHTVARWQRQFPFN', receivers, sub, filename)

    # 删除文件
    os.remove(filename)
