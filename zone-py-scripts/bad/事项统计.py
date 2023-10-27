import datetime
import gc
import os
import smtplib
import time
from collections import defaultdict
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pymongo
import pymysql
from openpyxl.workbook import Workbook

file_context = '事项关联情况统计'
env = 'DEV'
# env = 'PROD'
receivers = None
mongo_client = None
cursor = None

region_code_to_name_dict = {
    "0500": "苏州市",
    "0505": "高新区",
    "0506": "吴中区",
    "0507": "相城区",
    "0508": "姑苏区",
    "0509": "吴江区",
    "0581": "常熟市",
    "0582": "张家港市",
    "0583": "昆山市",
    "0585": "太仓市",
    "0000": "国家级",
    "1111": "省级"
}

def get_info_resource_status(status):
    if status == '0':
        return '已创建'
    elif status == '1':
        return '发布待审核'
    elif status == '2':
        return '发布已驳回'
    elif status == '3':
        return '已发布'
    elif status == '4':
        return '撤销待审核'
    elif status == '5':
        return '撤销已驳回'
    elif status == '6':
        return '已撤销'
    elif status == '7':
        return '已删除'
    return ''

def init_envs(env):
    if env == 'PROD':
        receivers = ['jason.lu@wingconn.com', 'tommie.chen@wingconn.com']
        mysql_conn = pymysql.connect(host="2.46.2.102", port=3306, user='bigdata', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient("2.46.2.239", 27017, username='admin', password='123@abcd',
                                           directConnection=True)
    else:

        receivers = ['jason.lu@wingconn.com']
        mysql_conn = pymysql.connect(host="mysql-master", port=3306, user='bigdata', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient(host="mongo-0.mongo,mongo-1.mongo,mongo-2.mongo`", port=27017)
        mongo_client.admin.authenticate('admin', '123@abcd', mechanism='SCRAM-SHA-1')
    # mysql 字典迭代器
    cursor = mysql_conn.cursor(cursor=pymysql.cursors.DictCursor)
    return mongo_client, cursor, receivers


# mysql查询 返回python dict array
def get_mysql_res(sql):
    cursor.execute(sql)
    res = cursor.fetchall()
    return res


# mongo查询 返回python dict array
def get_mongo_res(mongo_client, database, collection, query, projection=None):
    if projection is None:
        results = mongo_client[database][collection].find(query)
    else:
        results = mongo_client[database][collection].find(query, projection)
    return list(results)


### 发送邮件
def send_mail(smtp_host, port, sender, password, receivers, sub, filename):
    msg = MIMEMultipart()
    msg['Subject'] = sub
    msg['From'] = sender

    part = MIMEText('本邮件为自动发送，请勿回复！')
    msg.attach(part)

    part = MIMEApplication(open(filename, 'rb').read())
    part.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(part)

    smtp = smtplib.SMTP(smtp_host, port, 'utf-8')
    smtp.login(sender, password)
    smtp.sendmail(sender, receivers, msg.as_string())
    smtp.quit()
    print('send msg successfully!')


if __name__ == '__main__':
    mongo_client, cursor, receivers = init_envs(env)
    task_list = get_mysql_res(
        "SELECT t.task_guid as task_guid, t.task_name as task_name, t.task_type as task_type, m.dept_id, m.dept_name as dept_name FROM cmp_catalog.task t left join cmp_catalog.task_dept_map m on t.ou_code = m.ou_code")
    dept_name_to_task_list_map = defaultdict(list)
    for task in task_list:
        dept_name_to_task_list_map[task['dept_name']].append(task)
    info_res_list = get_mysql_res(
        "select id, info_resource_name, info_resource_provider,create_date,info_resource_life_cycle, is_open, regioncode, info_item_desc_detail, share_type, related_task_flag, task_type, task_name, basic_catalog_code, business_handling_code, implement_list_code from cmp_catalog.info_resource_platform")
    del task_list
    gc.collect()
    dept_name_to_info_res_list_map = defaultdict(list)
    for info_res in info_res_list:
        info_res['task_name_list'] = list(filter(None, (info_res.get('task_name') or "").split("|")))
        dept_name_to_info_res_list_map[info_res['info_resource_provider']].append(info_res)
    del info_res_list
    gc.collect()

    detail_sheet_row_list = []
    detail_sheet_row_list.append([
        '部门名称',
        '事项名称',
        '目录名称',
        '目录发布状态',
    ])
    count_shet_row_list = []
    count_shet_row_list.append([
        '部门名称',
        '事项数量',
        '已被信息资源关联数量',
        '未被信息资源关联数量',
    ])
    for dept_name in dept_name_to_task_list_map.keys():
        task_list = dept_name_to_task_list_map.get(dept_name, [])
        info_res_list = dept_name_to_info_res_list_map.get(dept_name, [])
        task_count = len(task_list)
        task_has_related_in_info_res_count = 0
        for task in task_list:
            task_name = task['task_name']
            task_has_related_in_info_res = False
            for info_res in info_res_list:
                info_res_related_task_name_list = info_res.get('task_name_list',[])
                for info_res_related_task in info_res_related_task_name_list:
                    if info_res_related_task == task_name:
                        task_has_related_in_info_res = True
                        detail_sheet_row_list.append([
                            dept_name,
                            task_name,
                            info_res['info_resource_name'],
                            get_info_resource_status(info_res['info_resource_life_cycle'])
                        ])

            if task_has_related_in_info_res:
                task_has_related_in_info_res_count = task_has_related_in_info_res_count+1
            else:
                detail_sheet_row_list.append([
                    dept_name,
                    task_name,
                    None,
                    None
                ])
        count_shet_row_list.append([
            dept_name,
            task_count,
            task_has_related_in_info_res_count,
            task_count-task_has_related_in_info_res_count
        ])

    filename = file_context + '.xlsx'
    # 创建一个工作簿
    wb = Workbook()

    # 获取活动工作表
    ws = wb.active
    ws.title = '部门事项统计'
    # 将数据写入工作表
    for row in count_shet_row_list:
        ws.append(row)
        print(row)
    # 将所有单元格格式设置为字符串
    for row in ws.iter_rows():
        for cell in row:
            cell.number_format = '@'

    ws2 = wb.create_sheet('详细情况')
    # 将数据写入工作表
    for row in detail_sheet_row_list:
        ws2.append(row)
        print(row)


    # 将所有单元格格式设置为字符串
    for row in ws2.iter_rows():
        for cell in row:
            cell.number_format = '@'

    # 将工作簿保存到 xlsx 文件
    wb.save(filename)

    ###
    sub = Header('[' + env + ']-[' + file_context + ']' + str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                 'utf-8')
    send_mail('smtp.163.com', 25, 'rikurobot@163.com', 'BDDHTVARWQRQFPFN', receivers, sub, filename)

    # 删除文件
    os.remove(filename)
