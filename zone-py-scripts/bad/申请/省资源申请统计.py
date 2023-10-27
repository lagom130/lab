import csv
import datetime
import gc
import itertools
import json
import os
import re
import smtplib
import time
import zipfile
from collections import defaultdict
from email.header import Header
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pymongo
import pymysql
from openpyxl.workbook import Workbook

file_context = '省资源申请统计'
env = '326'
# env = 'DEV'
# env = 'PROD'
receivers = None
mongo_client = None
cursor = None


#########################################################################################################
######## 字典 ###########################################################################################
# 属地字典
region_code_to_name_dict = {
    "0500": "苏州市",
    "0505": "高新区",
    "0506": "吴中区",
    "0507": "相城区",
    "0508": "姑苏区",
    "0509": "吴江区",
    "0581": "常熟市",
    "0582": "张家港市",
    "0583": "昆山市",
    "0585": "太仓市",
    "0590": "工业园区",
    "0000": "国家级",
    "1111": "省级"
}
city_region = '0500'
area_regions = ['0505', '0506', '0507', '0508', '0509', '0581', '0582', '0583', '0585', '0590']
province_regions = ['0000', '1111']
filter_info_resource_code_prefix = '30701505003030050000000003'

resuse_scenes_dict = {'1': '二级建造师执业资格认定', '10': '食品安全企业标准备案',
                      '100': '从事经营性互联网文化活动审批', '101': '其他', '11': '限制进出口货物的许可证审批',
                      '12': '二级建造师注册资格认定', '13': '注册造价工程师执业资格注册的核准',
                      '14': '分批次建设用地审查', '15': '律师执业、变更、注销许可',
                      '16': '设计企业资质核准（甲级及部分乙级除外）', '17': '提取住房公积金审批',
                      '18': '二级注册结构工程师执业资格认定', '19': '单独选址建设项目用地审查',
                      '2': '施工单位的主要负责人、项目负责人、专职安全生产管理人员安全任职资格的核准',
                      '20': '企业设立、变更、注销登记', '21': '医疗机构配制的制剂品种和制剂调剂审批',
                      '22': '药品生产企业许可', '23': '危险废物转移跨省审批',
                      '24': '法律援助律师、公职律师、公司律师工作证颁发', '25': '工程造价咨询企业资质的核准',
                      '26': '第二类、第三类医疗器械生产许可', '27': '医疗器械广告审批',
                      '28': '危险化学品生产企业安全生产许可', '29': '工程监理企业资质核准（综合、专业甲级除外）',
                      '3': '企业设立、变更、注销登记', '30': '基本养老保险个人账户（个人账户余额）一次性支付核定',
                      '31': '第二类医疗器械产品注册审批', '32': '建设项目压覆重要矿产资源审批',
                      '33': '乙类大型医用设备配置许可', '34': '住房公积金贷款审批',
                      '35': '涉及饮用水卫生安全的产品卫生许可', '36': '社会团体成立、变更、注销登记',
                      '37': '药品经营许可', '38': '工程质量检测机构资质的核准', '39': '公路超限运输许可',
                      '4': '施工单位的特种作业人员操作资格证书的核准', '40': '房地产开发企业资质核定（二级及以下）',
                      '41': '部分易制毒化学品和石墨类相关制品进出口许可', '42': '勘察企业资质核准（乙级）',
                      '43': '司法鉴定机构及其分支机构设立、变更、注销登记', '44': '工程造价咨询企业备案',
                      '45': '变更研制新药、生产药品已获证明文件及附件中载明事项补充申请审批',
                      '46': '检验检测机构资质认定', '47': '药品广告审批和备案',
                      '48': '研制过程中所需研究用对照药品一次性进口', '49': '货物自动进口许可',
                      '5': '建筑业企业资质核准（总承包特级、一级、部分二级及部分专业承包一级、二级除外）',
                      '50': '食品（含保健食品）生产许可', '51': '药品、医疗器械信息服务审批',
                      '52': '城乡规划编制单位资质的核准', '53': '林木采伐许可证核发',
                      '54': '医师执业注册（含外国医师来华短期行医许可，台湾地区医师在大陆短期行医许可，香港、澳门特别行政区医',
                      '55': '建设项目用地预审', '56': '永久性占用征收林地', '57': '二级注册建筑师执业资格认定',
                      '58': '5000万元以下的内资企业技术改造项目和3000万美元以下的外资企业技术改造项目进口设备免关税确认',
                      '59': '放射源诊疗技术和医用辐射机构许可', '6': '建筑施工企业安全生产许可证的核准',
                      '60': '勘察设计注册工程师执业资格认定', '61': '外国企业常驻代表机构登记（设立、变更、注销）',
                      '62': '医疗机构放射性职业病危害建设项目预评价报告审核', '63': '职业（工种）技能鉴定',
                      '64': '建设项目压覆重要矿床（矿产资源）审批', '65': '期刊、报纸变更刊期、报纸变更开版审批',
                      '66': '省外从事建设监理中介服务活动的机构进入本省承接业务资格（资质）的核验',
                      '67': '企业印制发票审批（增值税专用发票以外其他发票）', '68': '内部资料性出版物准印证核发',
                      '69': '药品委托生产审批', '7': '企业职工基本养老保险待遇审核', '70': '内部资料性出版物准印证核发',
                      '71': '律师事务所（分所）设立、变更、注销许可', '72': '武器装备科研生产单位二级、三级保密资格认定',
                      '73': '广告发布登记', '74': '非煤矿山企业安全生产许可',
                      '75': '工程建设中采用没有国家、行业及地方标准的新技术、新工艺、新材料的审定',
                      '76': '民办非企业单位成立、变更、注销登记', '77': '涉路施工活动许可',
                      '78': '药品生产质量管理规范（GMP）认证', '79': '气象计量器具检定', '8': '毕业生就业报到证的签发',
                      '80': '房地产开发企业一级资质核定', '81': '编印、发送宗教内部资料性出版物或者印刷其他宗教用品审批',
                      '82': '审批本行政区域内增值电信业务经营许可的申请', '83': '对纳税人延期缴纳税款的核准',
                      '84': '一级资质的房地产估价机构新设立分支机构的备案', '85': '基金会成立、变更、注销登记',
                      '86': '特种设备生产单位许可', '87': '建筑施工企业安全生产许可证核发', '88': '保健食品广告审批',
                      '89': '出售、购买、利用国家重点保护野生动物及其制品审批',
                      '9': '企业职工和离退休人员因病或非因工死亡及供养直系亲属待遇核定', '90': '辐射安全许可',
                      '91': '药品经营质量管理规范（GSP）认证',
                      '92': '金融机构营业场所、金库安全防范设施建设方案审批及工程验收', '93': '食品（含保健食品）生产许可',
                      '94': '公路建设项目竣工验收', '95': '超限高层建筑工程抗震设防审批',
                      '96': '律师事务所（分所）设立、变更、注销许可', '97': '探矿权延续登记',
                      '98': '企业、事业单位、社会团体等投资建设的固定资产投资项目核准', '99': '医疗机构执业登记'}

use_range_dict = {
    '1':'行政依据',
    '2':'工作参考',
    '3':'业务协同',
    '4':'数据校核',
    '5':'其他',
}
apply_basis_type_dict = {
    "1":"国家政策",
    "2":"省级政策",
    "3":"市级政策",
    "99":"其他",
}

# 目录分类（按行业）
belong_industry_dict = {
    "A": "农、林、牧、渔业",
    "B": "采矿业",
    "C": "制造业",
    "D": "电力、势力、燃气及水生产和供应业",
    "E": "建筑业",
    "F": "批发和零售业",
    "G": "交通运输、仓储和邮政业",
    "H": "住宿和餐饮业",
    "I": "信息传输、软件和信息技术服务业",
    "J": "金融业",
    "K": "房地产业",
    "L": "租赁和商务服务业",
    "M": "科学研究和技术服务业",
    "N": "水利、环境和公共设施管理业",
    "O": "居民服务、修理和其他服务业",
    "P": "教育",
    "Q": "卫生和社会工作",
    "R": "文化、体育和娱乐业",
    "S": "公共管理、社会保障和社会组织",
    "T": "国际组织"
}
# 布尔中文-字符串01
bool_number_string_dict = {
    "1": "是",
    "0": "否"
}
# 是否电子证照
certification_type_dict = {
    "1": "是",
    "2": "否"
}
# 事项目录所属领域
field_type_dict = {
    "1": "科技创新",
    "2": "商贸流通",
    "3": "社会救助",
    "4": "城建住房",
    "5": "教育文化",
    "6": "工业农业",
    "7": "机构团体",
    "8": "地理空间",
    "9": "资源能源",
    "10": "市场监管",
    "11": "生活服务",
    "12": "生态环境",
    "13": "交通运输",
    "14": "安全生产",
    "15": "社保就业",
    "16": "医疗卫生",
    "17": "信用服务",
    "18": "公共服务",
    "19": "财税金融",
    "20": "气象服务",
    "21": "法律服务",
    "22": "新冠疫苗",
    "23": "其他"
}
# 应用场景
gov_use_scene_dict = {
    "1": "政务服务",
    "2": "公共服务",
    "3": "监管",
    "0": "其他"
}
# 目录生命周期
info_res_life_cycle_dict = {
    "0": "已创建",
    "1": "发布待审核",
    "2": "发布已驳回",
    "3": "已发布",
    "4": "撤销待审核",
    "5": "撤销已驳回",
    "6": "已撤销",
    "7": "已删除"
}
# 信息项共享类型
item_share_type_dict = {
    "1": "无条件共享",
    "2": "有条件共享",
    "3": "不予共享"
}
# 提供渠道
net_type_dict = {
    "1": "政务外网",
    "2": "互联网",
    "3": "部门专网"
}
# 开放类型
open_type_dict = {
    "unconditional": "无条件开放",
    "conditional": "有条件开放",
    "disallowed": "不开放"
}
# 不xx依据
reason_type_dict = {
    1: "法律法规",
    2: "政策文件",
    3: "涉及敏感信息",
    4: "其他"
}
# 所属领域
resource_subject_dict = {
    "01": "婚育收养",
    "02": "环境资源",
    "03": "教育培训",
    "04": "住房保障",
    "05": "劳动就业",
    "06": "执业资格",
    "07": "社会保障",
    "08": "纳税缴费",
    "09": "医疗卫生",
    "10": "交通旅游",
    "11": "出境入境",
    "12": "生活服务",
    "13": "民族宗教",
    "14": "三农服务",
    "15": "死亡殡葬",
    "16": "设立变更",
    "17": "行业准营",
    "18": "投资立项",
    "19": "工程建设",
    "20": "涉外服务",
    "21": "安全生产",
    "22": "破产注销",
    "23": "文物保护",
    "24": "其他"
}
# 数据敏感级别
sensitive_level_dict = {
    "1": "1级",
    "2": "2级",
    "3": "3级",
    "4": "4级"
}
# 存储类型
storage_mode_dict = {
    "API": "接口",
    "TABLE": "数据库",
    "FILE_STRUCTURED": "结构化文件",
    "FILE_UNSTRUCTURED": "非结构化文件"
}
# 事项类型
task_type_dict = {
    "GOV": "政务服务事项",
    "PUBLIC": "公共服务事项"
}
# 目录分类（按主题）
use_scene_dict = {
    "01": "社会保险",
    "02": "交通出行",
    "03": "婚育",
    "04": "社区周边生活服务",
    "05": "政府办事",
    "06": "就医与保健",
    "07": "学校教育与终身教育",
    "08": "培训与就业",
    "09": "城市安全"
}
# 布尔值字典
bool_dict = {
    True: '是',
    False: '否'
}
# 布尔值字符串字典
bool_str_dict = {
    'true': '是',
    'false': '否'
}
# 布尔值字符串字典
bool_yn_dict = {
    'Y': '是',
    'N': '否'
}
# 更新周期
update_cycle_dict = {
    'year': '每年',
    'quarter': '季度',
    'month': '每月',
    'week': '每周',
    'day': '每日',
    'ondemand': '按需',
    '实时': '实时',
    '每日': '每日',
    '每周': '每周',
    '每月': '每月',
    '季度': '季度',
    '每年': '每年',
    '按需': '按需'
}

# 资源状态
data_res_status_dict = {
    'created': '已创建',
    'publishAudit': '发布待审核',
    'published': '已发布',
    'publishRejected': '发布已驳回',
    'revokeAudit': '撤销待审核',
    'revoked': '已撤销',
    'revokeRejected': '撤销已驳回',
    'closed': '已关闭',
}
# 匹配模式
url_map_mode_dict = {
    'prefix': '前缀匹配',
    'absolute': '绝对匹配',
}
# 请求类型
api_protocol_dict = {
    'rs': 'restful',
    'ws': 'webservice'
}
# 资源类型
resource_type_dict = {
    'API': '接口',
    'TABLE': '库表',
    'FILE': '文件',
    'api': '接口',
    'table': '库表',
    'file': '文件',
}
# 申请状态
apply_status_dict = {
    'created': '已创建',
    'unAudit': '待审核',
    'auditReject': '已退回',
    'unSubscrib': '待订阅',
    'unSend': '待推送',
    'finished': '已完成',
    'invalid': '失效',
    'handOver': '转交'
}
# 省级上报状态
report_status_dict = {
    None: '未发布',
    '': '未发布',
    'handled': '已绑定',
    'initializing': '初始化中',
    'initialized': '初始化完成',
    'migrating': '迁移中',
    'migrated': '迁移完成',
    'published': '已发布',
    'unpublished': '不发布'}
# 目录省级上报状态
info_resource_report_status_dict = {
    None: '未上报',
    '': '未上报',
    'success': '已上报'
}

###########################################################################################################
######## 通用方法 ###########################################################################################

# 根据字典key获取字典值，缺省默认值为空字符串
def dict_key_to_value(dictionary, key, default_value=""):
    return dictionary.get(key, default_value)


# 根据字典值获取字典key，缺省默认值为空字符串
def dict_value_to_key(dictionary, value, default_value=""):
    for k, v in dictionary.items():
        if v == value:
            return k
    return default_value


def get_element(arr, i, default_value=""):
    try:
        return arr[i]
    except IndexError:
        return default_value


# 单一实体占用多行写入单元格，只有第一行会写
def write_cell(entity_index, value):
    if entity_index == 0:
        return value
    return None


# 单一实体占用多行写入单元格，只有第一行会写，并将值翻译
def write_cell_convert_by_dict(entity_index, dictionary, key, default_value=""):
    if entity_index == 0:
        return dict_key_to_value(dictionary, key, default_value)
    return None


# 单一实体占用多行写入单元格，只有第一行会写
def entity_to_cell(entity_index, entity, key):
    if entity_index == 0:
        return entity.get(key, None)
    return None


# 单一实体占用多行写入单元格，只有第一行会写，并将值翻译
def entity_to_cell_convert_by_dict(entity_index, dictionary, entity, key, default_value=""):
    if entity_index == 0:
        return dict_key_to_value(dictionary, entity.get(key, default_value), default_value)
    return None


# 初始化环境信息
def init_envs(env):
    if env == 'PROD':
        receivers = ['jason.lu@wingconn.com', 'tommie.chen@wingconn.com', 'sisi.zhong@wingconn.com', 'xiao.liu@wingconn.com','emily.yuan@wingconn.com']
        mysql_conn = pymysql.connect(host="2.46.2.102", port=3306, user='bigdata', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient("2.46.2.239", 27017, username='admin', password='123@abcd',
                                           directConnection=True)
    elif env == '326':
        receivers = ['jason.lu@wingconn.com']
        mysql_conn = pymysql.connect(host="10.10.32.6", port=3306, user='wingtest', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient("10.10.32.6", port=27017)
        mongo_client.admin.authenticate('admin', '123@abcd', mechanism='SCRAM-SHA-1')
    else:

        receivers = ['jason.lu@wingconn.com']
        mysql_conn = pymysql.connect(host="mysql-master", port=3306, user='bigdata', password='123@abcd',
                                     charset='utf8mb4')
        mongo_client = pymongo.MongoClient(host="mongo-0.mongo,mongo-1.mongo,mongo-2.mongo`", port=27017)
        mongo_client.admin.authenticate('admin', '123@abcd', mechanism='SCRAM-SHA-1')
    # mysql 字典迭代器
    cursor = mysql_conn.cursor(cursor=pymysql.cursors.DictCursor)
    return mongo_client, cursor, receivers


# 13位时间戳转格式化日期字符串
def timestamp_to_date_str(time_stamp, format_string="%Y-%m-%d %H:%M:%S"):
    try:
        if time_stamp is None or time_stamp == '':
            return ''
        if len(str(time_stamp)) > 13:
            time_stamp = int(str(time_stamp)[:13])
        elif len(str(time_stamp)) < 13:
            time_stamp = int('{}{}'.format(str(time_stamp), '0' * (10 - len(str(time_stamp)))))

        time_array = time.localtime((time_stamp / 1000))
        return time.strftime(format_string, time_array)
    except BaseException as e:
        print(str(time_stamp))
        return ''


# mysql查询 返回python dict array
def get_mysql_res(sql):
    cursor.execute(sql)
    res = cursor.fetchall()
    return res


# mongo查询 返回python dict array
def get_mongo_res(mongo_client, database, collection, query, projection=None):
    if projection is None:
        results = mongo_client[database][collection].find(query)
    else:
        results = mongo_client[database][collection].find(query, projection)
    return list(results)


def write_data_to_sheet(sheet, sheet_name, titles, data_list):
    sheet.title = sheet_name
    sheet.append(titles)
    if data_list is not None and len(data_list) > 0:
        for row in data_list:
            sheet.append(row)
    for row in sheet.iter_rows():
        for cell in row:
            cell.number_format = '@'


def write_csv_to_zip(path, title, data_list, zip_file):
    with open(path, 'w', encoding='utf-8', newline='') as cf:
        writer = csv.writer(cf)
        writer.writerow(title)
        if data_list is not None:
            for item in data_list:
                for cell in item:
                    if cell is None or cell is 'null':
                        cell = ''
                writer.writerow(item)
    zip_file.write(path)
    os.remove(path)


### 发送邮件
def send_mail(smtp_host, port, sender, password, receivers, sub, filename):
    msg = MIMEMultipart()
    msg['Subject'] = sub
    msg['From'] = sender

    part = MIMEText('本邮件为自动发送，请勿回复！')
    msg.attach(part)

    part = MIMEApplication(open(filename, 'rb').read())
    part.add_header('Content-Disposition', 'attachment', filename=filename)
    msg.attach(part)

    smtp = smtplib.SMTP(smtp_host, port, 'utf-8')
    smtp.login(sender, password)
    smtp.sendmail(sender, receivers, msg.as_string())
    smtp.quit()
    print('send msg successfully!')


# 列表转字符串（sql拼接用）
def get_arr_str(arr):
    return "'" + "','".join(arr) + "'"

if __name__ == '__main__':
    mongo_client, cursor, receivers = init_envs(env)
    applies = get_mongo_res(mongo_client, 'data_share_db', 'data_resource_apply',
                            {'data_creater_dep_id':'admin'},
                            {'_id': 1, 'resource_id': 1, 'share_resource_id': 1, 'resource_name': 1, 'apply_type':1,'code': 1, 'status': 1,
                             'dep_name': 1, 'resource_name': 1 ,'data_creater_dep_name':1, 'apply_time': 1,
                             "updated_time":1, "use_range":1, "resuse_scenes":1, "use_desc":1,
                             "apply_basis_type":1,"apply_basis_content":1,"origin_incident":1})
    operation_records = get_mongo_res(mongo_client, 'data_share_db', 'operation_record',
                                      {'data_type': 'shareResourceApply'},
                                      {'data_id': 1, 'data_name': 1, 'data_status': 1, 'created_time': 1})
    operation_records_dict = defaultdict(list)
    for item in operation_records:
        operation_records_dict[item['data_id']].append(item)
    for key in operation_records_dict:
        operation_records_dict[key] = sorted(operation_records_dict[key], key=lambda x: x['created_time'], reverse=True)
    row_list = []
    row_list.append(['申请主键','resourceId','shareResourceId','申请流水号',
                     '申请方','资源名称','资源类型','申请状态','申请时间','更新时间',
                     '资源使用范围','资源使用场景','资源用途说明','资源申请依据','资源申请依据内容',
                     '申请使用事由'])
    for apply in applies:
        resource_name = apply.get('resource_name', '')
        apply_type = apply.get('apply_type', '')
        if apply_type == 'API' and resource_name.startswith('区县级联'):
            continue
        status = dict_key_to_value(apply_status_dict,apply.get('status', ''), '')
        records = operation_records_dict.get(str(apply['_id']), [])
        if status == '失效':
            ceng_finished = False
            ceng_un_subscrib = False
            for item in records:
                if 'finished' in item.get('data_status',''):
                    ceng_finished = True
                if 'unSubscrib' in item.get('data_status', ''):
                    ceng_un_subscrib = True
            if ceng_finished:
                status = status + '(已完成)'
            elif ceng_un_subscrib:
                status = status = '(审核通过)'
        resuse_scenes = apply.get('resuse_scenes', '').split(',')
        resuse_scenes_value_arr = []
        resuse_scenes_values = ''
        if len(resuse_scenes) > 0:
            for item in resuse_scenes:
                resuse_scenes_value_arr.append(dict_key_to_value(resuse_scenes_dict, item, ''))
            resuse_scenes_values = ",".join(resuse_scenes_value_arr)
        row_list.append([
            str(apply['_id']),
            str(apply.get('resource_id','')),
            str(apply.get('share_resource_id','')),
            apply.get('code', ''),
            apply.get('dep_name', ''),
            resource_name,
            dict_key_to_value(resource_type_dict, apply_type, ''),
            status,
            timestamp_to_date_str(apply.get('apply_time', '')),
            apply.get('updated_time', ''),
            dict_key_to_value(use_range_dict,apply.get('use_range', ''), apply.get('use_range', '')),
            resuse_scenes_values,
            apply.get('use_desc', ''),
            dict_key_to_value(apply_basis_type_dict, apply.get('apply_basis_type', ''), ''),
            apply.get('apply_basis_content', ''),
            apply.get('origin_incident', ''),
        ])

    filename = file_context + '.xlsx'
    # 创建一个工作簿
    wb = Workbook()

    # 获取活动工作表
    ws = wb.active
    ws.title = file_context
    # 将数据写入工作表
    for row in row_list:
        ws.append(row)
        print(row)
    # 将所有单元格格式设置为字符串
    for row in ws.iter_rows():
        for cell in row:
            cell.number_format = '@'


    # 将工作簿保存到 xlsx 文件
    wb.save(filename)

    ###
    sub = Header('[' + env + ']-[' + file_context + ']' + str(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                 'utf-8')
    send_mail('smtp.163.com', 25, 'rikurobot@163.com', 'BDDHTVARWQRQFPFN', receivers, sub, filename)

    # 删除文件
    os.remove(filename)

    print('export task ended at:' + datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
